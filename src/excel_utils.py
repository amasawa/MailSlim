"""
Excel import/export utilities for email candidate lists.
"""

import openpyxl
from pathlib import Path

HEADERS = ["Approve (Y/N)", "Folder", "From", "From Name", "Subject", "Date", "Reason", "UID"]


def export_candidates(candidates, output_path):
    """Export deletion candidates to Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Delete Candidates"
    ws.append(HEADERS)

    for c in candidates:
        ws.append([
            c["approve"], c["folder"], c["from"], c["from_name"],
            c["subject"], c["date"], c["reason"], c["uid"]
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(output_path)
    print(f"Exported {len(candidates)} candidates to: {output_path}")


def load_approved(approved_path):
    """Load approved deletion list from Excel. Returns list of dicts with uid, subject, folder."""
    wb = openpyxl.load_workbook(approved_path)
    ws = wb.active

    to_delete = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        approve, folder, frm, frm_name, subject, date, reason, uid = row
        if str(approve).strip().upper() == "Y" and uid:
            to_delete.append({"uid": uid, "subject": subject, "folder": folder})
    return to_delete
